# File Name:   FileReader.py
# Author:      Hao Ye
# Version:     V1.0 09/04/2018
# Version:     V2.0 11/12/2018
# Purpose:     Transform coordinate data from OSGB coordinate to a local coordinate system #
# Usage:       No arguments needed, but GriderBuilder.py must be in same directory         #

import os
print "Local Projection Transformation Version 1.0"
print "OSTN15 -> Local Projection "
print "\nLoading Modules... "
import arcpy
import openpyxl
import glob

###############################################################################################
class Coordinate(object):
    def __init__(self, zoneKey, id, coordE, coordN, height):
        '''Read from constructor'''
        self.ZoneKey = zoneKey          # Zone Key parameter
        self.FID = id
        self.CoordE = coordE            # Easting Coordinate
        self.CoordN = coordN            # Easting Coordinate
        self.Height = height            # Height Coordinate
        self.PrjSF = None               # Project Scale Factor
        self.ElvSF = None               # Elevation Scale Factor
        self.CSF = None                 # Combined Scale Factor       
        self.StartCordE = None
        self.StartCordN = None    
        self.OrignCordE = None
        self.OrignCordN = None       
        self.LclCordE = None
        self.LclCordN = None

    def calcuCSF(self):
        if self.PrjSF != None and self.ElvSF != None:
            self.CSF = self.PrjSF * self.ElvSF
        else:
            print '\nThe factor is null'

    def calcuESCoord(self):
        if self.CSF != None:
            self.LclCordE = (self.CoordE - self.OrignCordE)/(self.CSF)
            self.LclCordN = (self.CoordN - self.OrignCordN)/(self.CSF)
        else:
            print '\nThe factor is null'

    def BuildGeometry(self, point):
        newPoint = arcpy.Point()
        newPoint.X = self.LclCordE
        newPoint.Y = self.LclCordN
        pointGeometry = arcpy.PointGeometry(newPoint)
        return pointGeometry

    @property
    def setPrjSF(self):
        return self.PrjSF

    @setPrjSF.setter
    def setPrjSF(self, value):
        self.PrjSF = value

    @property
    def getZoneKey(self):
        return self.ZoneKey

    @property
    def getHeight(self):
        return self.Height


class Zones(object):  
    def __init__(self, key, startE, startN, orignE, orignN, prjSF):
        self.Key = key
        self.StartE = startE            # Loop local grid parameters
        self.StartN = startN            # Loop local grid parameters
        self.OrignE = orignE
        self.OrignN = orignN
        self.PrjSF = prjSF

    @property
    def getPrjSF(self):
        return self.PrjSF

    @property
    def getKey(self):
        return self.Key
    

class HetBand(object):
    def __init__(self, height, hegtBand, elvSF, comts):
        self.Height= height
        self.HegtBand = hegtBand
        self.ElvSF = elvSF
        self.Comts = comts

# #############################################################################

# Define the project folder
projectPath = os.path.dirname(os.path.dirname(os.path.abspath(sys.argv[0])))
workspace = os.path.join(projectPath,'WorkingSpace')
resultfd = os.path.join(projectPath,'Result')
confspace = os.path.join(projectPath,'DevelopersOnly')
arcpy.env.workspace = workspace
print "\nProgram Starts, your current working folder is " + projectPath

# Find the shapefile under the folder
for file in os.listdir(workspace):
    if file.endswith('.shp') or file.endswith('.SHP'):
        fc = os.path.join(workspace,file)

# Find the configuration under the folder
for file in os.listdir(confspace):
    if file.endswith('.xlsx'):
        cg = os.path.join(confspace,file)
      
resultfile =  os.path.join(resultfd, "ProjectLayer.shp")
of = os.path.join(resultfd, "ProjectLayer")

# Set up essential variables required for shapefile processing
fields = ['FID','EASTING','NORTHING','COVER_LEVE']                  # Fields extracted from shapefile
checfd = ['LocalEast', 'LocalNorth']                                # Checked existing Fields
instfd = ['FID','LocalEast', 'LocalNorth']                          # Fields needed to be inserted

# Configure lists to config grids, height band and points
listPoints = []                                                     # Candindate points from shapefile  
listAreas = []
listLocalGrid = []                                                  # Local grids reference
listHetBand = []                                                    # Height band reference
pntGemtry = []                                                      # List of point geometries


# Config the zones for local grid system
wb = openpyxl.load_workbook(cg)                                     # Read excel file to retrieve grid parameters
sheet = wb['Grid']
for row in range(2, sheet.max_row + 1):   
    zkey = sheet['A'+str(row)].value                                # Zone Value
    staE = sheet['B'+str(row)].value                                # Northing Value
    staN = sheet['C'+str(row)].value                                # Easting Value
    orgE = sheet['D'+str(row)].value                                # Height Value
    orgN = sheet['E'+str(row)].value                                # Height Value
    prjF = sheet['F'+str(row)].value                                # Projection SF Value 
    zoneCell = Zones(zkey, staE, staN, orgE,orgN, prjF)             # Zone Value
    listLocalGrid.append(zoneCell)                                  # Build Grid Value0

# Config the Height bands for local grid system
sheet = wb['Band']
for row in range(2, sheet.max_row + 1):   
    hegt = sheet['A'+str(row)].value                                # Height Value
    elSF = sheet['B'+str(row)].value                                # Northing Value
    heBd = sheet['C'+str(row)].value                                # Elevation Scale Factor
    comt = sheet['D'+str(row)].value                                # Comments
    bandCell = HetBand(hegt, elSF, heBd, comt)                      # Zone Value
    listHetBand.append(bandCell) 

# Configure the areas
sheet = wb['Grid']
for row in range(2, sheet.max_row + 1):   
    area = sheet['A'+str(row)].value
    listAreas.append(area)

print "\nProgram Configured, data processing starts"

while True:
    zoneKey = raw_input('\nPlease type the Zone Key, valid values between A1-A30, B15-B32, C13-C19: ')
    if zoneKey in listAreas:
        print "\nArea is in the list, You agree to process the data"
        time.sleep(1)
        break
    else:
        print "\nArea is not in the list, please reset up the data"
        time.sleep(1)
        exit()

print "\nData is processing,please wait."

# Extract field data from from shpfile
cursor = arcpy.da.SearchCursor (fc, fields)                         # Read customised field from script start
for row in cursor:
    coords = Coordinate(zoneKey, row[0], row[1], row[2], row[3])
    listPoints.append(coords)
del cursor

# Populate the parameters to the points list
for coord in listPoints:                                            # Check points in the list
    for zone in listLocalGrid:                                      # Check zones in local grid
        if coord.ZoneKey == zone.getKey:                            # Make sure key is same
            coord.setPrjSF = zone.getPrjSF                          # Projection SF
            coord.StartCordE = zone.StartE                          # Zone Start E
            coord.StartCordN = zone.StartN                          # Zone Start N
            coord.OrignCordE = zone.OrignE                          # Zone Origin E
            coord.OrignCordN = zone.OrignN                          # Zone Origin N
 
# Determine average height of points of the site
sum = 0.0
for coord in listPoints:
    sum += coord.Height
aveHeight = sum/len(listPoints)

# Deterimine the height band of points and pupulate band index
for coord in listPoints:
    coord.ElvSF = 0.9999875
    coord.calcuCSF()            #Pupulate CSF values
    coord.calcuESCoord()        #Populate East and North Value                                

# Check if the coordinate fields have been created
fieldList = arcpy.ListFields(fc)
for field in fieldList:
    existingField = False
    if field.name == "LocalEast" or field.name == "LocalNorth":
        existingField = True
    else:
        existingField = False

if existingField == False:
    arcpy.AddField_management(fc,"LocalEast","DOUBLE",15)
    arcpy.AddField_management(fc,"LocalNorth","DOUBLE",15)

# Copy all geometres from shapefile to new shapefile
for coord in listPoints:
    uc = arcpy.da.UpdateCursor(fc, instfd)
    for row in uc:
        if coord.FID==row[0]:
            print coord.FID, coord.LclCordE, coord.LclCordN
            newPt = [coord.FID,str(coord.LclCordE),str(coord.LclCordN)]
            uc.updateRow(newPt)
    del uc

# Create a new shapefile with transformed geometries
for coord in listPoints:
    geometry = coord.BuildGeometry(coord)
    pntGemtry.append(geometry)

# Check if the shapefile is existing, if yes update shapefile
if os.path.isfile(resultfile):
    files = glob.glob(resultfd + '/*')
    for f in files:
        os.remove(f)
    arcpy.CopyFeatures_management(pntGemtry,of)
else:
    arcpy.CopyFeatures_management(pntGemtry,of)

#for field in fieldList:
    #if field.name != 'FID' and field.name != 'Shape':
        #arcpy.AddField_management(resultfile, field.name,field.type,field.precision,field.scale,field.length,field.aliasName,field.isNullable,field.required,field.domain)

arcpy.MakeFeatureLayer_management(fc, "pk")
arcpy.JoinField_management(resultfile,'FID',"pk",'FID')
arcpy.DeleteField_management(resultfile,'id')

print "\nProgram Completed, please, check data in result folder"